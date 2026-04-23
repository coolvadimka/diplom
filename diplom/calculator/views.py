from django.shortcuts import render
from django.http import HttpResponse
from .forms import CalculatorForm
from .logic import KLogicCalculator, UnsafeExpressionError
from openpyxl import Workbook
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


MAX_ROWS = 100000  # максимальное количество строк k^n
PREVIEW_ROWS = 20 #максималльно количество строк выводимое на странице
MAX_FORM_ROWS = 81

def format_expression_error(error: Exception) -> str:  # переводим внутренние исключения в понятные сообщения для пользователя
    if isinstance(error, SyntaxError):  # если сам Python-парсер не смог разобрать выражение
        return "Ошибка синтаксиса"  # оставляем короткое стандартное сообщение

    if isinstance(error, UnsafeExpressionError):  # отдельно обрабатываем наши контролируемые ошибки валидации и парсинга
        message = str(error)  # достаём исходный текст ошибки из исключения

        if message.startswith("Константа ") or message == "Константа должна быть целым числом":  # сообщения про неверные константы уже достаточно конкретные
            return message  # показываем их пользователю без упрощения

        if (  # перечисляем новые сообщения парсера, которые важно не прятать за общим текстом
            message == "Формула не задана"  # пользователь ничего не ввёл
            or message == "Не закрыта скобка"  # функция или выражение осталось без закрывающей скобки
            or message == "Ожидалась закрывающая скобка"  # не найдена закрывающая скобка для подвыражения
            or message == "Ожидалось выражение"  # выражение оборвалось после оператора или в другом недопустимом месте
            or message.startswith("После символа ¬")  # ошибка аргумента после отрицания Поста
            or message.startswith("После символа ~")  # ошибка аргумента после отрицания Лукашевича
            or message.startswith("Недопустимый символ:")  # в выражении встретился запрещённый символ
            or message.startswith("Ожидалась ',' или ')'")  # ошибка в списке аргументов функции
            or message.startswith("Ошибка синтаксиса возле:")  # парсер смог показать место, где сломался разбор
        ):
            return message  # возвращаем точное сообщение, чтобы пользователь и ты при тестировании видели реальную причину

        return "Ошибка синтаксиса"  # остальные ошибки уровня синтаксиса сводим к общему сообщению

    if isinstance(error, ValueError):  # ошибки бизнес-логики, например неверное количество аргументов
        return str(error)  # показываем как есть, потому что эти тексты уже человекочитаемые

    if isinstance(error, TypeError):  # ошибки типов значений
        return str(error)  # тоже не скрываем, потому что они могут быть полезны при проверках

    return "Ошибка при вычислении выражения"  # общий запасной вариант на случай непредвиденного исключения

def calculator_view(request):
    basic_forms = None
    forms_error = None
    form = CalculatorForm(request.POST or None)
    variables = None
    table = None
    error = None
    total_rows = 0
    preview_rows = PREVIEW_ROWS
    is_truncated = False
    result_name = "F"

    if request.method == "POST" and form.is_valid():
        k = form.cleaned_data["k"]
        n = form.cleaned_data["n"]
        expression = form.cleaned_data["expression"].strip()

        try:
            calc = KLogicCalculator(k)

            calc.validate_expression(expression)

            detected_variables = calc.extract_variables(expression)
            row_count = k ** len(detected_variables) if detected_variables else 1

            if row_count > MAX_ROWS:
                error = (
                    f"Слишком большая таблица: {row_count} строк. "
                    f"Уменьши k или количество переменных в формуле (лимит {MAX_ROWS})."
                )
            else:
                variables, full_table, _ = calc.generate_table(expression, n)
                result_name = expression
                total_rows = len(full_table)

                if total_rows <= MAX_FORM_ROWS:
                    basic_forms = calc.build_basic_forms(expression, n)
                else:
                    forms_error = (
                        f"Основные формы доступны только для таблиц размером до {MAX_FORM_ROWS} строк."
                    )

                if total_rows > PREVIEW_ROWS:
                    table = full_table[:PREVIEW_ROWS]
                    is_truncated = True
                else:
                    table = full_table
                    is_truncated = False

        except (ValueError, UnsafeExpressionError, SyntaxError, TypeError) as e:
            error = format_expression_error(e)
        except Exception:
            error = "Не удалось вычислить выражение. Проверь синтаксис."

    return render(request, "calculator/index.html", {
        "form": form,
        "variables": variables,
        "table": table,
        "error": error,
        "total_rows": total_rows,
        "preview_rows": preview_rows,
        "is_truncated": is_truncated,
        "result_name": result_name,
        "basic_forms": basic_forms,
        "forms_error": forms_error,
    })


def export_excel_view(request):
    k = request.GET.get("k")
    n = request.GET.get("n")
    expression = request.GET.get("expression", "").strip()
    if not k or not n or not expression:
        return HttpResponse("Не переданы параметры для экспорта.", status =400)
    try:
        k = int(k)
        n = int(n)

        calc = KLogicCalculator(k)
        variables = calc.extract_variables(expression)
        row_count = k ** len(variables) if variables else 1

        if row_count > MAX_ROWS:
            return HttpResponse(
                f"Слишком большая таблица для экспорта: {row_count} строк.",
                status=400
            )
        variables, full_table, _ = calc.generate_table(expression, n)
        result_name = expression

        wb = Workbook()
        ws = wb.active
        ws.title = "Truth table"

        # -------- Стили --------
        title_fill = PatternFill("solid", fgColor="1F4E78")
        section_fill = PatternFill("solid", fgColor="D9EAF7")
        header_fill = PatternFill("solid", fgColor="DCE6F1")
        result_fill = PatternFill("solid", fgColor="EAF2F8")

        title_font = Font(bold=True, color="FFFFFF", size=14)
        section_font = Font(bold=True, color="1F1F1F", size=11)
        label_font = Font(bold=True, color="1F1F1F")
        value_font = Font(color="1F1F1F")
        table_header_font = Font(bold=True, color="17365D")
        result_font = Font(bold=True, color="154D8C")

        thin_gray = Side(style="thin", color="D9D9D9")
        bottom_border = Border(bottom=thin_gray)

        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")

        # -------- Заголовок --------
        headers = variables + [result_name]
        last_col = len(headers)
        table_header_row = 12

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
        ws["A1"] = "K-значный логический калькулятор"
        ws["A1"].fill = title_fill
        ws["A1"].font = title_font
        ws["A1"].alignment = left_align

        # -------- Паспорт расчёта --------
        ws["A3"] = "Формула:"
        ws["B3"] = expression

        ws["A4"] = "k:"
        ws["B4"] = k

        ws["A5"] = "n (лимит):"
        ws["B5"] = n

        ws["A6"] = "Переменные:"
        ws["B6"] = ", ".join(variables)

        ws["A7"] = "Количество строк:"
        ws["B7"] = len(full_table)

        ws["A8"] = "Дата генерации:"
        ws["B8"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        for cell_ref in ["A3", "A4", "A5", "A6", "A7", "A8"]:
            ws[cell_ref].font = label_font
            ws[cell_ref].fill = section_fill
            ws[cell_ref].alignment = left_align

        for cell_ref in ["B3", "B4", "B5", "B6", "B7", "B8"]:
            ws[cell_ref].font = value_font
            ws[cell_ref].alignment = left_align

        # -------- Подзаголовок таблицы --------
        ws.merge_cells(start_row=10, start_column=1, end_row=10, end_column=last_col)
        ws["A10"] = "Таблица истинности"
        ws["A10"].fill = section_fill
        ws["A10"].font = section_font
        ws["A10"].alignment = left_align


        for col_index, header in enumerate(headers, start=1):
            cell = ws.cell(row=table_header_row, column=col_index, value=header)
            cell.fill = header_fill
            cell.font = table_header_font
            cell.alignment = center_align
            cell.border = bottom_border

        # -------- Данные --------
        for row_index, row in enumerate(full_table, start=table_header_row + 1):
            values = row["values"] + [row["result"]]

            for col_index, value in enumerate(values, start=1):
                cell = ws.cell(row=row_index, column=col_index, value=value)
                cell.alignment = center_align

                # выделяем столбец результата F
                if col_index == len(headers):
                    cell.fill = result_fill
                    cell.font = result_font

        # -------- Ширины колонок --------
        for col_index in range(1, len(headers) + 1):
            max_length = 0
            column_letter = get_column_letter(col_index)

            for row_num in range(1, ws.max_row + 1):
                cell = ws.cell(row=row_num, column=col_index)
                cell_value = str(cell.value) if cell.value is not None else ""
                if len(cell_value) > max_length:
                    max_length = len(cell_value)

            ws.column_dimensions[column_letter].width = max(max_length + 2, 12)

        # Немного воздуха
        ws.row_dimensions[1].height = 24
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A13"
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="truth_table.xlsx"'

        wb.save(response)
        return response

    except (ValueError, UnsafeExpressionError, SyntaxError, TypeError) as e:
        return HttpResponse(f"Ошибка экспорта: {e}", status=400)
    except Exception:
        return HttpResponse("Не удалось сформировать Excel-файл.", status=500)